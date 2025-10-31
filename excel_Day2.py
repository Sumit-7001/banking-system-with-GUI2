import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment

BANKING_ACCOUNT = "ACCOUNT"
ADMIN_LOGS = "ADMIN_LOGS"

if not os.path.exists(ADMIN_LOGS):
    os.makedirs(ADMIN_LOGS)

if not os.path.exists(BANKING_ACCOUNT):
    os.makedirs(BANKING_ACCOUNT)



def GetAccount(acc_no):
    return os.path.join(BANKING_ACCOUNT, f"Account__{acc_no}.xlsx")

def get_balance(file_path):
    wb = load_workbook(file_path)
    ws = wb["Account"]
    value = ws["A9"].value   
    wb.close()
    if value and ":" in value:
        try:
            return float(value.split(":")[1].strip())
        except ValueError:
            return 0.0
    return 0.0

def update_balance(file_path, new_balance):
    wb = load_workbook(file_path)
    ws = wb["Account"]
    ws["A9"] = f"BALANCE : {new_balance}"
    ws["A9"].alignment = Alignment(horizontal="center", vertical="center")
    wb.save(file_path)
    wb.close()

def save_transaction(file_path, t_type, amount, new_balance):
    wb = load_workbook(file_path)
    ws = wb["Account"]
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
   
    ws.append([now, t_type, amount, new_balance])
    wb.save(file_path)
    wb.close()


def create_new_account(acc_no, name, balance, aadhar_no, mail_id, dob, address):
    file_path = GetAccount(acc_no)
    if os.path.exists(file_path):
        print("Account already exists!")
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = "Account"

    ws.merge_cells('A1:D1')  
    ws['A1'] = "BANK NAME : BRAINWARE BAKWAS BANK"
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")

    ws["A2"] = "Account Holder"
    ws["B2"] = name
    ws["A3"] = "Account Number"
    ws["B3"] = acc_no
    ws["A4"] = "Aadhar NO"
    ws["B4"] = aadhar_no
    ws["A5"] = "Mail ID"
    ws["B5"] = mail_id
    ws["A6"] = "DOB"
    ws["B6"] = dob
    ws["A7"] = "Address"
    ws["B7"] = address
    
    ws.merge_cells('A9:D9') 
    ws["A9"] = f"BALANCE : {balance}"
    ws["A9"].alignment = Alignment(horizontal="center", vertical="center")

   
    ws["A11"] = "Date"
    ws["B11"] = "Type"
    ws["C11"] = "Amount"
    ws["D11"] = "New Balance"
    # -----------------------------------

    wb.save(file_path)
    return True

def get_last_transactions(file_path, num=5):
    wb = load_workbook(file_path)
    ws = wb["Account"]
    transactions = []
    
   
    HEADER_ROW = 11 
    
    for row in range(ws.max_row, HEADER_ROW, -1):
    # -----------------------------------
        if len(transactions) >= num:
            break 
            
        date = ws.cell(row=row, column=1).value
        ttype = ws.cell(row=row, column=2).value
        amount = ws.cell(row=row, column=3).value
        new_balance = ws.cell(row=row, column=4).value
        
        if date: 
            transactions.append({
                "date": date,
                "type": ttype,
                "amount": amount,
                "balance": new_balance
            })
            
    wb.close()
    return transactions


def GetAdminLogFile(admin_username):
    

    return os.path.join(ADMIN_LOGS, f"{admin_username}_log.xlsx")

def _init_admin_log_file(file_path, admin_username):
    

    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "ActivityLog"
       
        headers = ["Timestamp", "Admin", "Action", "Account Affected", "Amount"]
        ws.append(headers)
        wb.save(file_path)
        wb.close()

def save_admin_transaction(admin_username, action_type, account_no, amount):

    file_path = GetAdminLogFile(admin_username)
    _init_admin_log_file(file_path, admin_username)

    wb = load_workbook(file_path)
    ws = wb.active
    
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_entry = [now, admin_username, action_type, account_no, amount]
    ws.append(log_entry)
    
    wb.save(file_path)
    wb.close()

def get_all_recent_activities(admin_username, num=5):
    file_path = GetAdminLogFile(admin_username)
    
    if not os.path.exists(file_path):
        return [] 

    wb = load_workbook(file_path)
    ws = wb.active
    activities = []
    
    
    for row in range(ws.max_row, 1, -1):
        if len(activities) >= num:
            break 
            
        date = ws.cell(row=row, column=1).value
        admin = ws.cell(row=row, column=2).value
        action = ws.cell(row=row, column=3).value
        account_no = ws.cell(row=row, column=4).value
        amount = ws.cell(row=row, column=5).value
        
        if date: 
            activities.append({
                "date": date,
                "admin": admin,
                "action": action,
                "account_no": account_no,
                "amount": amount
            })
            
    wb.close()
    return activities


def get_account_details(file_path):
    

    wb = load_workbook(file_path)
    ws = wb["Account"]
    details = {
        "name": ws["B2"].value,     
        "acc_no": ws["B3"].value     
    }
    wb.close()
    return details

def get_all_transactions(file_path):
  

    wb = load_workbook(file_path)
    ws = wb["Account"]
    transactions = []
    HEADER_ROW = 11 
    
    for row in range(ws.max_row, HEADER_ROW, -1):
        
        date = ws.cell(row=row, column=1).value
        ttype = ws.cell(row=row, column=2).value
        amount = ws.cell(row=row, column=3).value
        new_balance = ws.cell(row=row, column=4).value
        
        if date: 
            transactions.append({
                "date": date,
                "type": ttype,
                "amount": amount,
                "balance": new_balance
            })
            
    wb.close()
    return transactions 